from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.services.document_processing_exceptions import XlsxTextExtractionError

EXTRACTION_METHOD = "openpyxl"


@dataclass(frozen=True, slots=True)
class ExtractedWorksheet:
    worksheet_index: int
    name: str
    text: str


@dataclass(frozen=True, slots=True)
class XlsxTextExtractionResult:
    worksheets: tuple[ExtractedWorksheet, ...]
    full_text: str
    worksheet_count: int


def extract_xlsx_text(content: bytes) -> XlsxTextExtractionResult:
    """Extract worksheet names and cell contents from an XLSX file."""
    if not content:
        raise XlsxTextExtractionError("XLSX file is empty")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise XlsxTextExtractionError("Failed to open XLSX for text extraction") from exc

    try:
        worksheets: list[ExtractedWorksheet] = []
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            row_lines = _extract_worksheet_rows(worksheet)
            readable_text = _format_worksheet_text(worksheet.title, row_lines)
            worksheets.append(
                ExtractedWorksheet(
                    worksheet_index=index,
                    name=worksheet.title,
                    text=readable_text,
                ),
            )

        full_text = "\n\n".join(sheet.text for sheet in worksheets if sheet.text)
        return XlsxTextExtractionResult(
            worksheets=tuple(worksheets),
            full_text=full_text,
            worksheet_count=len(worksheets),
        )
    finally:
        workbook.close()


def _extract_worksheet_rows(worksheet) -> list[str]:
    rows: list[str] = []
    for row in worksheet.iter_rows(values_only=True):
        cells = [_format_cell_value(value) for value in row]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(" | ".join(cells))
    return rows


def _format_cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_worksheet_text(name: str, rows: list[str]) -> str:
    if not rows:
        return f"[{name}]"

    body = "\n".join(rows)
    return f"[{name}]\n{body}"
