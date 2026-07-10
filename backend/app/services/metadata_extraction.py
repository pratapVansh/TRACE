"""Extract structural and document properties from uploaded files."""

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
import re

import fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image
from pptx import Presentation

from app.services.document_processing_exceptions import MetadataExtractionError

FILE_METADATA_KEY = "file_metadata"
PDF_DATE_PATTERN = re.compile(
    r"^D:(?P<year>\d{4})(?P<month>\d{2})?(?P<day>\d{2})?"
    r"(?P<hour>\d{2})?(?P<minute>\d{2})?(?P<second>\d{2})?",
)


@dataclass(frozen=True, slots=True)
class DocumentFileMetadata:
    page_count: int | None
    file_size_bytes: int
    file_type: str
    file_extension: str
    author: str | None
    creation_date: str | None
    modification_date: str | None

    def to_storage_dict(self) -> dict[str, object]:
        return {
            "page_count": self.page_count,
            "file_size_bytes": self.file_size_bytes,
            "file_type": self.file_type,
            "file_extension": self.file_extension,
            "author": self.author,
            "creation_date": self.creation_date,
            "modification_date": self.modification_date,
        }


def extract_document_metadata(
    content: bytes,
    *,
    mime_type: str,
    file_extension: str,
    existing_page_count: int | None = None,
) -> DocumentFileMetadata:
    """Extract file metadata for any supported upload type."""
    if not content:
        raise MetadataExtractionError("File is empty")

    extension = file_extension.lower().lstrip(".")
    file_size_bytes = len(content)

    try:
        extracted = _extract_by_extension(content, extension)
    except MetadataExtractionError:
        raise
    except Exception as exc:
        raise MetadataExtractionError(
            f"Failed to extract metadata for .{extension} file",
        ) from exc

    page_count = existing_page_count if existing_page_count is not None else extracted.page_count

    return DocumentFileMetadata(
        page_count=page_count,
        file_size_bytes=file_size_bytes,
        file_type=mime_type,
        file_extension=extension,
        author=extracted.author,
        creation_date=extracted.creation_date,
        modification_date=extracted.modification_date,
    )


def _extract_by_extension(content: bytes, extension: str) -> DocumentFileMetadata:
    if extension == "pdf":
        return _extract_pdf_metadata(content)
    if extension == "docx":
        return _extract_docx_metadata(content)
    if extension == "pptx":
        return _extract_pptx_metadata(content)
    if extension == "xlsx":
        return _extract_xlsx_metadata(content)
    if extension in {"png", "jpg", "jpeg"}:
        return _extract_image_metadata(content, extension)
    if extension == "txt":
        return DocumentFileMetadata(
            page_count=1,
            file_size_bytes=len(content),
            file_type="text/plain",
            file_extension=extension,
            author=None,
            creation_date=None,
            modification_date=None,
        )

    return DocumentFileMetadata(
        page_count=None,
        file_size_bytes=len(content),
        file_type="application/octet-stream",
        file_extension=extension,
        author=None,
        creation_date=None,
        modification_date=None,
    )


def _extract_pdf_metadata(content: bytes) -> DocumentFileMetadata:
    document = fitz.open(stream=content, filetype="pdf")
    try:
        meta = document.metadata or {}
        return DocumentFileMetadata(
            page_count=document.page_count,
            file_size_bytes=len(content),
            file_type="application/pdf",
            file_extension="pdf",
            author=_clean_text(meta.get("author")),
            creation_date=_parse_pdf_date(meta.get("creationDate")),
            modification_date=_parse_pdf_date(meta.get("modDate")),
        )
    finally:
        document.close()


def _extract_docx_metadata(content: bytes) -> DocumentFileMetadata:
    document = DocxDocument(BytesIO(content))
    properties = document.core_properties
    return DocumentFileMetadata(
        page_count=None,
        file_size_bytes=len(content),
        file_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        file_extension="docx",
        author=_clean_text(properties.author),
        creation_date=_format_datetime(properties.created),
        modification_date=_format_datetime(properties.modified),
    )


def _extract_pptx_metadata(content: bytes) -> DocumentFileMetadata:
    presentation = Presentation(BytesIO(content))
    properties = presentation.core_properties
    return DocumentFileMetadata(
        page_count=len(presentation.slides),
        file_size_bytes=len(content),
        file_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        file_extension="pptx",
        author=_clean_text(properties.author),
        creation_date=_format_datetime(properties.created),
        modification_date=_format_datetime(properties.modified),
    )


def _extract_xlsx_metadata(content: bytes) -> DocumentFileMetadata:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        properties = workbook.properties
        return DocumentFileMetadata(
            page_count=len(workbook.sheetnames),
            file_size_bytes=len(content),
            file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_extension="xlsx",
            author=_clean_text(properties.creator),
            creation_date=_format_datetime(properties.created),
            modification_date=_format_datetime(properties.modified),
        )
    finally:
        workbook.close()


def _extract_image_metadata(content: bytes, extension: str) -> DocumentFileMetadata:
    image = Image.open(BytesIO(content))
    try:
        author = None
        creation_date = None
        modification_date = None

        exif = image.getexif()
        if exif:
            for tag_id, value in exif.items():
                tag_name = Image.ExifTags.TAGS.get(tag_id, "")
                if tag_name == "Artist" and value:
                    author = _clean_text(str(value))
                elif tag_name in {"DateTime", "DateTimeOriginal"} and value:
                    parsed = _parse_exif_datetime(str(value))
                    creation_date = creation_date or parsed
                    modification_date = modification_date or parsed

        mime_type = f"image/{'jpeg' if extension in {'jpg', 'jpeg'} else extension}"
        return DocumentFileMetadata(
            page_count=1,
            file_size_bytes=len(content),
            file_type=mime_type,
            file_extension=extension,
            author=author,
            creation_date=creation_date,
            modification_date=modification_date,
        )
    finally:
        image.close()


def _clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped or None


def _format_datetime(value: datetime | None) -> str | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.isoformat()


def _parse_pdf_date(value: str | None) -> str | None:
    if not value:
        return None

    match = PDF_DATE_PATTERN.match(value.strip())
    if not match:
        return _clean_text(value)

    parts = match.groupdict(default="01")
    try:
        parsed = datetime(
            int(parts["year"]),
            int(parts["month"] or "01"),
            int(parts["day"] or "01"),
            int(parts["hour"] or "00"),
            int(parts["minute"] or "00"),
            int(parts["second"] or "00"),
            tzinfo=UTC,
        )
    except ValueError:
        return _clean_text(value)

    return parsed.isoformat()


def _parse_exif_datetime(value: str) -> str | None:
    try:
        parsed = datetime.strptime(value.strip(), "%Y:%m:%d %H:%M:%S").replace(tzinfo=UTC)
    except ValueError:
        return _clean_text(value)
    return parsed.isoformat()
