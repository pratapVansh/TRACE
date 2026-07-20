import os
import base64
import json
from typing import Any

from app.agents.framework.tool import ToolResult
from app.agents.framework.tools.base import FrameworkTool
from app.agents.framework.tools.context import ToolContext
from app.agents.framework.tools.schemas import ToolCategory, ToolMetadata
from app.core.authorization import Permission

def _get_safe_path(filename: str) -> str:
    """Ensure the path is strictly inside the workspace directory."""
    from app.core.config import settings
    base = str(settings.workspace_root_path.resolve())
    target = os.path.abspath(os.path.join(base, filename))
    if not target.startswith(base):
        raise ValueError(f"Access denied: Path '{filename}' is outside the workspace.")
    return target


class WorkspaceListTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="workspace_list",
        name="Workspace List",
        description="Lists files in the persistent workspace.",
        category=ToolCategory.ACTION,
        permissions={Permission.WORKSPACE},
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Subdirectory to list"}
            }
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        try:
            target_dir = _get_safe_path(params.get("path", ""))
            if not os.path.isdir(target_dir):
                return ToolResult(data=None, error="Not a directory")
            items = []
            for item in os.listdir(target_dir):
                item_path = os.path.join(target_dir, item)
                items.append({
                    "name": item,
                    "is_dir": os.path.isdir(item_path),
                    "size_bytes": os.path.getsize(item_path) if os.path.isfile(item_path) else 0
                })
            context.add_reasoning_step("WorkspaceListTool: listed directory")
            return ToolResult(data={"items": items})
        except Exception as e:
            return ToolResult(data=None, error=str(e))


class WorkspaceReadTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="workspace_read",
        name="Workspace Read",
        description="Reads files from the workspace. Supports extraction for some formats.",
        category=ToolCategory.ACTION,
        permissions={Permission.WORKSPACE},
        input_schema={
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "File to read"},
                "extract_text": {"type": "boolean", "description": "Whether to extract text from binary"}
            },
            "required": ["filename"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        try:
            filename = params["filename"]
            target = _get_safe_path(filename)
            if not os.path.isfile(target):
                return ToolResult(data=None, error="File not found")
            ext = os.path.splitext(target)[1].lower()
            extract_text = params.get("extract_text", False)
            if extract_text:
                if ext == ".pdf":
                    import fitz
                    text = "\n".join(page.get_text() for page in fitz.open(target))
                    return ToolResult(data={"filename": filename, "content": text})
                elif ext == ".xlsx":
                    import openpyxl
                    wb = openpyxl.load_workbook(target, data_only=True)
                    lines = []
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            lines.append("\t".join(str(c) if c else "" for c in row))
                    return ToolResult(data={"filename": filename, "content": "\n".join(lines)})
            
            text_exts = [".txt", ".md", ".csv", ".json", ".py"]
            if ext in text_exts or extract_text:
                with open(target, "r", encoding="utf-8") as f:
                    content = f.read()
                return ToolResult(data={"filename": filename, "content": content})
            else:
                with open(target, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                return ToolResult(data={"filename": filename, "content_base64": b64})
        except Exception as e:
            return ToolResult(data=None, error=str(e))


class WorkspaceWriteTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="workspace_write",
        name="Workspace Write",
        description="Writes a file. Can generate Excel files from JSON arrays automatically.",
        category=ToolCategory.ACTION,
        permissions={Permission.WORKSPACE},
        input_schema={
            "type": "object",
            "properties": {
                "filename": {"type": "string"},
                "content": {"type": "string"},
                "is_base64": {"type": "boolean"}
            },
            "required": ["filename", "content"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        try:
            filename = params["filename"]
            content = params["content"]
            is_b64 = params.get("is_base64", False)
            target = _get_safe_path(filename)
            ext = os.path.splitext(target)[1].lower()
            
            if is_b64:
                with open(target, "wb") as f:
                    f.write(base64.b64decode(content))
                return ToolResult(data={"success": True})
            
            if ext == ".xlsx" and not is_b64:
                import openpyxl
                try:
                    data = json.loads(content)
                    if isinstance(data, list) and data and isinstance(data[0], dict):
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        headers = list(data[0].keys())
                        ws.append(headers)
                        for r in data:
                            ws.append([r.get(h, "") for h in headers])
                        wb.save(target)
                        return ToolResult(data={"success": True})
                except Exception:
                    pass
            
            with open(target, "w", encoding="utf-8") as f:
                f.write(content)
            return ToolResult(data={"success": True})
        except Exception as e:
            return ToolResult(data=None, error=str(e))


class WorkspaceDeleteTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="workspace_delete",
        name="Workspace Delete",
        description="Deletes a file from the workspace.",
        category=ToolCategory.ACTION,
        permissions={Permission.WORKSPACE},
        input_schema={
            "type": "object",
            "properties": {
                "filename": {"type": "string"}
            },
            "required": ["filename"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        try:
            target = _get_safe_path(params["filename"])
            if os.path.exists(target):
                if os.path.isdir(target):
                    import shutil
                    shutil.rmtree(target)
                else:
                    os.remove(target)
                return ToolResult(data={"success": True})
            return ToolResult(data=None, error="Not found")
        except Exception as e:
            return ToolResult(data=None, error=str(e))
