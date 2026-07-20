import csv
import io
import json
from typing import Any

from sqlalchemy import text
from app.db.session import engine

from app.agents.framework.tool import ToolResult
from app.agents.framework.tools.base import FrameworkTool
from app.agents.framework.tools.context import ToolContext
from app.agents.framework.tools.schemas import ToolCategory, ToolMetadata
from app.core.authorization.permissions import Permission


class SqlTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="sql_execute",
        name="SQL Execution",
        description="Executes a read-only SQL query against the connected database.",
        category=ToolCategory.ACTION,
        permissions={Permission.AI_AGENTS},
        input_schema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The SQL query to execute"}
            },
            "required": ["query"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        query = params.get("query", "")
        if not query.strip():
            return ToolResult(data=None, error="Query cannot be empty")
        
        # Simple safety check (naive)
        if any(kw in query.upper() for kw in ["DROP", "DELETE", "UPDATE", "INSERT", "ALTER", "TRUNCATE", "GRANT", "REVOKE"]):
            return ToolResult(data=None, error="Only SELECT queries are allowed")

        try:
            async with engine.connect() as conn:
                result = await conn.execute(text(query))
                keys = list(result.keys())
                rows = [dict(zip(keys, row)) for row in result.all()]
                context.add_reasoning_step(f"SqlTool: executed query, retrieved {len(rows)} rows.")
                return ToolResult(data={"rows": rows, "count": len(rows)})
        except Exception as e:
            return ToolResult(data=None, error=f"SQL execution failed: {e}")


class CsvTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="csv_processor",
        name="CSV Processor",
        description="Reads and parses a CSV string into a structured JSON array.",
        category=ToolCategory.DATA,
        permissions=set(),
        input_schema={
            "type": "object",
            "properties": {
                "csv_data": {"type": "string", "description": "Raw CSV string data"},
                "delimiter": {"type": "string", "description": "Delimiter (default is comma)"}
            },
            "required": ["csv_data"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        csv_data = params.get("csv_data", "")
        delimiter = params.get("delimiter", ",")
        
        if not csv_data.strip():
            return ToolResult(data=None, error="csv_data is empty")

        try:
            f = io.StringIO(csv_data)
            reader = csv.DictReader(f, delimiter=delimiter)
            rows = [row for row in reader]
            context.add_reasoning_step(f"CsvTool: parsed {len(rows)} rows from CSV.")
            return ToolResult(data={"rows": rows})
        except Exception as e:
            return ToolResult(data=None, error=f"CSV processing failed: {e}")


class ExcelTool(FrameworkTool):
    metadata = ToolMetadata(
        tool_id="excel_processor",
        name="Excel Processor",
        description="Reads data from a base64 encoded Excel file.",
        category=ToolCategory.DATA,
        permissions=set(),
        input_schema={
            "type": "object",
            "properties": {
                "file_b64": {"type": "string", "description": "Base64 encoded Excel file"},
                "sheet_name": {"type": "string", "description": "Target sheet name (optional)"}
            },
            "required": ["file_b64"]
        }
    )

    async def execute(self, params: dict[str, Any], context: ToolContext) -> ToolResult:
        import base64
        import openpyxl
        
        file_b64 = params.get("file_b64", "")
        sheet_name = params.get("sheet_name")

        if not file_b64:
            return ToolResult(data=None, error="file_b64 is empty")

        try:
            file_bytes = base64.b64decode(file_b64)
            f = io.BytesIO(file_bytes)
            wb = openpyxl.load_workbook(f, data_only=True)
            
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([str(c) if c is not None else "" for c in row])
                
            context.add_reasoning_step(f"ExcelTool: parsed {len(data)} rows from sheet {ws.title}.")
            return ToolResult(data={"sheet_name": ws.title, "rows": data})
        except Exception as e:
            return ToolResult(data=None, error=f"Excel processing failed: {e}")
