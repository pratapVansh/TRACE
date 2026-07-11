from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.logging import logger
from app.models.document import Document
from app.models.document_version import DocumentVersion
from app.processing.base import BaseProcessor

SHEET_SEPARATOR = "\n\n" + "=" * 48 + "\n"
TABLE_SEPARATOR = "-" * 40


class ExcelProcessor(BaseProcessor):
    name = "excel_processor"
    supported_extensions = frozenset({"xlsx"})
    supported_mime_types = frozenset({
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })

    async def extract_text(
        self,
        document: Document,
        version: DocumentVersion,
    ) -> str:
        file_path = self._resolve_path(version)
        logger.info("Opening XLSX document_id=%s path=%s", document.id, file_path)

        try:
            wb = load_workbook(
                file_path,
                data_only=False,
                read_only=False,
                keep_vba=False,
            )
        except Exception as exc:
            logger.error("Failed to open XLSX document_id=%s: %s", document.id, exc)
            return ""

        try:
            return self._extract_all_sheets(wb)
        finally:
            wb.close()

    def _extract_all_sheets(self, wb) -> str:
        parts: list[str] = []

        for ws in wb.worksheets:
            parts.append(SHEET_SEPARATOR)
            state_label = " (hidden)" if ws.sheet_state == "hidden" else ""
            parts.append(f"Worksheet: {ws.title}{state_label}")
            parts.append(SHEET_SEPARATOR)

            self._extract_sheet_content(ws, parts)

        result = "\n".join(parts).strip()
        return result

    def _extract_sheet_content(self, ws, parts: list[str]) -> None:
        dim = ws.dimensions
        parts.append(f"Range: {dim}")
        parts.append("")

        if ws.max_row is None or ws.max_row == 0:
            parts.append("(empty sheet)")
            parts.append("")
            return

        merged_ranges = list(ws.merged_cells.ranges)

        for row_idx in range(1, ws.max_row + 1):
            row_parts: list[str] = []

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)
                coord = f"{col_letter}{row_idx}"

                is_merged = any(coord in m for m in merged_ranges)

                if is_merged:
                    is_top_left = any(
                        m.min_row == row_idx and m.min_col == col_idx
                        for m in merged_ranges
                    )

                    if not is_top_left:
                        continue

                value = cell.value
                if value is None:
                    continue

                display = self._format_cell_value(coord, value)
                if display:
                    row_parts.append(display)

            if row_parts:
                parts.extend(row_parts)
                parts.append("")

        if ws.tables:
            parts.append("")
            parts.append(TABLE_SEPARATOR)
            parts.append("Tables:")
            for table_name in ws.tables:
                table_obj = ws.tables[table_name]
                parts.append(f"  - {table_obj.displayName}: {table_obj.ref}")
            parts.append(TABLE_SEPARATOR)
            parts.append("")

    def _format_cell_value(self, coord: str, value) -> str | None:
        if isinstance(value, str) and value.startswith("="):
            return f"  {coord}: [FORMULA] {value}"

        if isinstance(value, str):
            return f"  {coord}: {value}"

        if isinstance(value, (int, float)):
            return f"  {coord}: {value}"

        if isinstance(value, datetime):
            return f"  {coord}: {value.isoformat()}"

        if isinstance(value, bool):
            return f"  {coord}: {str(value)}"

        return f"  {coord}: {value}"

    async def extract_metadata(
        self,
        document: Document,
        version: DocumentVersion,
    ) -> dict:
        file_path = self._resolve_path(version)

        try:
            wb = load_workbook(
                file_path,
                data_only=False,
                read_only=False,
                keep_vba=False,
            )
        except Exception as exc:
            logger.error("Failed to open XLSX for metadata document_id=%s: %s", document.id, exc)
            return self._fallback_metadata(document, version, file_path)

        try:
            return self._collect_metadata(wb, document, version, file_path)
        finally:
            wb.close()

    def _collect_metadata(
        self,
        wb,
        document: Document,
        version: DocumentVersion,
        file_path: str,
    ) -> dict:
        cp = wb.properties
        worksheets = wb.worksheets
        ws_count = len(worksheets)
        visible = 0
        hidden = 0
        total_rows = 0
        total_cols = 0
        formula_count = 0
        merged_cell_count = 0
        comment_count = 0
        hyperlink_count = 0
        table_count = 0
        image_count = 0
        chart_count = 0
        has_pivot = False

        for ws in worksheets:
            if ws.sheet_state == "hidden":
                hidden += 1
            else:
                visible += 1

            if ws.max_row:
                total_rows += ws.max_row
            if ws.max_column:
                total_cols += ws.max_column

            merged_cell_count += len(ws.merged_cells.ranges)
            table_count += len(ws.tables)

            if hasattr(ws, "_images"):
                image_count += len(ws._images)
            if hasattr(ws, "_charts"):
                chart_count += len(ws._charts)
            if hasattr(ws, "_pivots") and ws._pivots:
                has_pivot = True

            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        comment_count += 1
                    if cell.hyperlink:
                        hyperlink_count += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1

        created = self._format_datetime(cp.created)
        modified = self._format_datetime(cp.modified)

        has_macros = self._detect_macros(file_path)
        has_external_links = self._detect_external_links(file_path)

        result = {
            "title": cp.title or document.title,
            "author": cp.creator or "",
            "created": created,
            "modified": modified,
            "worksheet_count": ws_count,
            "visible_sheet_count": visible,
            "hidden_sheet_count": hidden,
            "row_count": total_rows,
            "column_count": total_cols,
            "formula_count": formula_count,
            "merged_cell_count": merged_cell_count,
            "comment_count": comment_count,
            "hyperlink_count": hyperlink_count,
            "table_count": table_count,
            "image_count": image_count,
            "chart_count": chart_count,
            "file_size": self._get_file_size(file_path),
            "extension": "xlsx",
            "mime_type": version.mime_type,
        }

        flags = []
        if chart_count > 0:
            flags.append("chart")
            result["requires_chart_processing"] = True
        if image_count > 0:
            flags.append("image")
            result["requires_image_processing"] = True
        if has_pivot:
            flags.append("pivot")
            result["requires_pivot_processing"] = True
        if has_macros:
            flags.append("macro")
            result["has_macros"] = True
        if has_external_links:
            flags.append("external_link")
            result["has_external_links"] = True

        if flags:
            logger.info(
                "Detected in document_id=%s: %s", document.id, ", ".join(flags)
            )

        logger.info(
            "Metadata extracted document_id=%s sheets=%d formulas=%d comments=%d hyperlinks=%d tables=%d",
            document.id,
            ws_count,
            formula_count,
            comment_count,
            hyperlink_count,
            table_count,
        )
        return result

    async def validate(
        self,
        document: Document,
        version: DocumentVersion,
    ) -> list[str]:
        warnings: list[str] = []
        file_path = self._resolve_path(version)

        if self._is_encrypted_zip(file_path):
            warnings.append("Password-protected or encrypted XLSX file")
            return warnings

        try:
            wb = load_workbook(
                file_path,
                data_only=False,
                read_only=False,
                keep_vba=False,
            )
        except BadZipFile:
            warnings.append("Corrupted or invalid XLSX file (not a valid ZIP archive)")
            return warnings
        except Exception as exc:
            err_str = str(exc).lower()
            if "password" in err_str or "encrypted" in err_str:
                warnings.append("Password-protected or encrypted XLSX file")
            else:
                warnings.append(f"Cannot open XLSX: {exc}")
            return warnings

        try:
            if len(wb.worksheets) == 0:
                warnings.append("XLSX workbook is empty (no worksheets)")
                return warnings

            all_empty = True
            for ws in wb.worksheets:
                if ws.max_row and ws.max_row > 0 and ws.max_column and ws.max_column > 0:
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                all_empty = False
                                break
                        if not all_empty:
                            break
                if not all_empty:
                    break
            if all_empty:
                warnings.append("XLSX workbook contains no data")
        finally:
            wb.close()

        return warnings

    async def process(
        self,
        document: Document,
        version: DocumentVersion,
    ) -> "ProcessingResult":
        from app.processing.models import ProcessingResult

        warnings = await self.validate(document, version)
        if warnings:
            metadata = await self.extract_metadata(document, version)
            return ProcessingResult(
                success=True,
                document_id=document.id,
                extracted_text="",
                metadata=metadata,
                warnings=warnings,
            )

        metadata = await self.extract_metadata(document, version)
        text = await self.extract_text(document, version)

        warnings_list: list[str] = []
        if metadata.get("requires_chart_processing"):
            warnings_list.append("Charts detected. Chart data extraction not yet implemented.")
        if metadata.get("requires_image_processing"):
            warnings_list.append("Images detected. Image processing not yet implemented.")
        if metadata.get("has_macros"):
            warnings_list.append("Macros detected in workbook.")
        if metadata.get("has_external_links"):
            warnings_list.append("External links detected in workbook.")

        return ProcessingResult(
            success=True,
            document_id=document.id,
            extracted_text=text,
            metadata=metadata,
            warnings=warnings_list,
        )

    def _resolve_path(self, version: DocumentVersion) -> str:
        return str(settings.storage_root_path / version.storage_uri)

    @staticmethod
    def _is_encrypted_zip(file_path: str) -> bool:
        try:
            with ZipFile(file_path, "r") as zf:
                names = [n.lower() for n in zf.namelist()]
                for name in names:
                    if "encryptioninfo" in name:
                        return True
            return False
        except Exception:
            return False

    @staticmethod
    def _detect_macros(file_path: str) -> bool:
        try:
            with ZipFile(file_path, "r") as zf:
                names = [n.lower() for n in zf.namelist()]
                for name in names:
                    if "vba" in name or "vbaProject" in name or "macro" in name:
                        return True
            return False
        except Exception:
            return False

    @staticmethod
    def _detect_external_links(file_path: str) -> bool:
        try:
            with ZipFile(file_path, "r") as zf:
                names = [n.lower() for n in zf.namelist()]
                for name in names:
                    if "externallink" in name or "externalLink" in name:
                        return True
            return False
        except Exception:
            return False

    @staticmethod
    def _format_datetime(dt: datetime | None) -> str | None:
        if dt is None:
            return None
        try:
            return dt.isoformat()
        except (ValueError, AttributeError):
            return None

    @staticmethod
    def _get_file_size(file_path: str) -> int | None:
        try:
            return Path(file_path).stat().st_size
        except OSError:
            return None

    @staticmethod
    def _fallback_metadata(
        document: Document,
        version: DocumentVersion,
        file_path: str,
    ) -> dict:
        return {
            "title": document.title,
            "author": None,
            "worksheet_count": 0,
            "visible_sheet_count": 0,
            "hidden_sheet_count": 0,
            "row_count": 0,
            "column_count": 0,
            "formula_count": 0,
            "merged_cell_count": 0,
            "comment_count": 0,
            "hyperlink_count": 0,
            "table_count": 0,
            "image_count": 0,
            "chart_count": 0,
            "file_size": ExcelProcessor._get_file_size(file_path),
            "extension": "xlsx",
            "mime_type": version.mime_type,
        }
