"""Create comprehensive test XLSX files for ExcelProcessor validation."""

import pathlib
from datetime import datetime, UTC
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table

test_dir = pathlib.Path("tests/fixtures/xlsx")
test_dir.mkdir(parents=True, exist_ok=True)


def create_empty_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(str(test_dir / "empty.xlsx"))
    wb.close()
    print("Created empty.xlsx")


def create_single_sheet_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = 95
    ws["A3"] = "Bob"
    ws["B3"] = 25
    ws["C3"] = 87
    wb.save(str(test_dir / "single_sheet.xlsx"))
    wb.close()
    print("Created single_sheet.xlsx")


def create_multi_sheet_xlsx():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Total"
    ws1["B1"] = 100

    ws2 = wb.create_sheet("Details")
    ws2["A1"] = "Item"
    ws2["B1"] = "Qty"
    ws2["A2"] = "Widget"
    ws2["B2"] = 50
    ws2["A3"] = "Gadget"
    ws2["B3"] = 30

    ws3 = wb.create_sheet("HiddenReport")
    ws3.sheet_state = "hidden"
    ws3["A1"] = "Hidden Data"
    ws3["B1"] = "Secret"

    wb.save(str(test_dir / "multi_sheet.xlsx"))
    wb.close()
    print("Created multi_sheet.xlsx")


def create_formula_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = "Value1"
    ws["B1"] = "Value2"
    ws["C1"] = "Sum"
    ws["A2"] = 10
    ws["B2"] = 20
    ws["C2"] = "=A2+B2"
    ws["A3"] = 5
    ws["B3"] = 15
    ws["C3"] = "=A3*B3"
    wb.save(str(test_dir / "formula.xlsx"))
    wb.close()
    print("Created formula.xlsx")


def create_merged_cells_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged Header Row"
    ws["A2"] = "Col A"
    ws["B2"] = "Col B"
    ws["C2"] = "Col C"
    ws["A3"] = "Data1"
    ws["B3"] = "Data2"
    ws["C3"] = "Data3"
    wb.save(str(test_dir / "merged_cells.xlsx"))
    wb.close()
    print("Created merged_cells.xlsx")


def create_comments_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws["A1"] = "Cell with comment"
    ws["A1"].comment = Comment("This is an important cell", "Reviewer")
    ws["B1"] = "Another cell"
    ws["B1"].comment = Comment("Double-check this value", "QA Team")
    ws["C1"] = "No comment"
    wb.save(str(test_dir / "comments.xlsx"))
    wb.close()
    print("Created comments.xlsx")


def create_hyperlinks_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    ws["A1"] = "Click here"
    ws["A1"].hyperlink = "http://example.com"
    ws["A2"] = "Documentation"
    ws["A2"].hyperlink = "http://docs.example.com"
    ws["A3"] = "No link"
    wb.save(str(test_dir / "hyperlinks.xlsx"))
    wb.close()
    print("Created hyperlinks.xlsx")


def create_tables_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws["A1"] = "Product"
    ws["B1"] = "Quantity"
    ws["C1"] = "Price"
    ws["A2"] = "Widget A"
    ws["B2"] = 100
    ws["C2"] = 9.99
    ws["A3"] = "Widget B"
    ws["B3"] = 50
    ws["C3"] = 14.99

    tab = Table(displayName="InventoryTable", ref="A1:C3")
    ws.add_table(tab)
    wb.save(str(test_dir / "tables.xlsx"))
    wb.close()
    print("Created tables.xlsx")


def create_all_features_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Master"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Master Report"

    headers = ["ID", "Product", "Quantity", "Notes"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)

    data = [
        [1, "Item A", 100, "In stock"],
        [2, "Item B", 50, "=C3*2"],
        [3, "Item C", 200, "Backorder"],
    ]
    for ri, row_data in enumerate(data, 3):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)

    ws["A3"].comment = Comment("First item", "Manager")
    ws["B3"].hyperlink = "http://item-a.example.com"

    tab = Table(displayName="MasterTable", ref="A2:D5")
    ws.add_table(tab)

    ws2 = wb.create_sheet("Config")
    ws2.sheet_state = "hidden"
    ws2["A1"] = "Internal"
    ws2["B1"] = "Settings"

    wb.save(str(test_dir / "all_features.xlsx"))
    wb.close()
    print("Created all_features.xlsx")


def create_corrupted_xlsx():
    path = test_dir / "corrupted.xlsx"
    with open(str(path), "wb") as f:
        f.write(b"This is not a valid XLSX file\nGarbage content")
    print("Created corrupted.xlsx")


def create_password_protected_xlsx():
    path = test_dir / "password_protected.xlsx"

    single = test_dir / "single_sheet.xlsx"
    if not single.exists():
        create_single_sheet_xlsx()

    import shutil
    shutil.copy(str(single), str(path))

    with ZipFile(str(path), "a", ZIP_DEFLATED) as zf:
        zf.writestr("EncryptionInfo", b"<encryption><p:encrypted/></encryption>")
    print("Created password_protected.xlsx")


def create_large_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Large"

    for ci in range(1, 21):
        ws.cell(row=1, column=ci, value=f"Col{ci}")

    for ri in range(2, 102):
        for ci in range(1, 21):
            ws.cell(row=ri, column=ci, value=f"R{ri}C{ci}")

    wb.save(str(test_dir / "large.xlsx"))
    wb.close()
    print("Created large.xlsx")


if __name__ == "__main__":
    create_empty_xlsx()
    create_single_sheet_xlsx()
    create_multi_sheet_xlsx()
    create_formula_xlsx()
    create_merged_cells_xlsx()
    create_comments_xlsx()
    create_hyperlinks_xlsx()
    create_tables_xlsx()
    create_all_features_xlsx()
    create_corrupted_xlsx()
    create_password_protected_xlsx()
    create_large_xlsx()
    print("\nAll test XLSX files created successfully.")
